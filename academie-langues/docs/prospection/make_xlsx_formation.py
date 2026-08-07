# -*- coding: utf-8 -*-
"""Génère nexa-prospects-formation-elargie.xlsx — écoles, CFP, indépendants (pas seulement langues)."""
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import re

HEADERS = [
    "NOM ETS",
    "DATE",
    "NO TELEPHONE",
    "TYPE DE STRUCTURE",
    "REFERENCE",
    "RETOUR APRES  RELANCE",
    "VILLE",
]

REFERENCE = "ZIDANE"
DATE_DEFAUT = "03/08/2026"

# (nom, type, telephone, ville_code)
# Sources: sites officiels, GoAfricaOnline, Ayila'a — contacts publics
ROWS = [
    # --- Formation pro / CFP ---
    ["CENTRE DE FORMATION IGBS", DATE_DEFAUT, "681037698", "CENTRE DE FORMATION PRO", REFERENCE, "", "DLA"],
    ["CFP KAYLANG", DATE_DEFAUT, "680392096/ 658647800", "CENTRE DE FORMATION PRO", REFERENCE, "", "DLA"],
    ["DOUALA INSTITUTE TECHNOLOGY", DATE_DEFAUT, "233475974/ 677977902", "CENTRE DE FORMATION PRO", REFERENCE, "", "DLA"],
    ["CFPC LA SALLE (DLS)", DATE_DEFAUT, "233438359/ 696458863", "CENTRE DE FORMATION PRO", REFERENCE, "", "DLA"],
    ["CFPD DEUTOU", DATE_DEFAUT, "243055582/ 675131320", "CENTRE DE FORMATION PRO", REFERENCE, "", "DLA"],
    ["CFPE-AVTC DOUALA", DATE_DEFAUT, "697610299/ 677653741", "CENTRE DE FORMATION PRO", REFERENCE, "", "DLA"],
    ["CFP-DELCAR (ENGINS LOURDS)", DATE_DEFAUT, "691100141/ 673272477", "CENTRE DE FORMATION PRO", REFERENCE, "", "YDÉ, DLA, BAF"],
    ["CFPROHRA BAFOUSSAM", DATE_DEFAUT, "696806473", "CENTRE DE FORMATION PRO", REFERENCE, "", "BAF"],
    ["CIA FORMATION", DATE_DEFAUT, "692175099/ 680550775", "CENTRE DE FORMATION DIGITALE", REFERENCE, "", "YDÉ"],
    ["CFP LA VOCATION", DATE_DEFAUT, "657491158", "CENTRE DE FORMATION PRO", REFERENCE, "", "YDÉ"],
    ["CFP-NETFOSA", DATE_DEFAUT, "699872079/ 679808609", "CENTRE DE FORMATION PRO", REFERENCE, "", "YDÉ"],
    ["ACERFI FORMATION", DATE_DEFAUT, "695080808/ 672060060", "CENTRE DE FORMATION DIGITALE", REFERENCE, "", "YDÉ, DLA"],
    ["TELEC-FORMATION", DATE_DEFAUT, "222239281/ 657816491/ 671868635", "CENTRE DE FORMATION DIGITALE", REFERENCE, "", "YDÉ"],
    ["POWERBACHE EDUCATION", DATE_DEFAUT, "653518317", "CENTRE DE FORMATION PRO", REFERENCE, "", "YDÉ"],
    ["IFPKM SOA", DATE_DEFAUT, "670488835", "CENTRE DE FORMATION PRO", REFERENCE, "", "YDÉ"],
    ["CIS FORMATION DOUALA", DATE_DEFAUT, "698482438/ 672072984", "CENTRE DE FORMATION DIGITALE", REFERENCE, "", "DLA"],
    ["CIS FORMATION YAOUNDE", DATE_DEFAUT, "671049520/ 691914908", "CENTRE DE FORMATION DIGITALE", REFERENCE, "", "YDÉ"],
    ["MULTIPURPOSE CENTER MBOUO", DATE_DEFAUT, "699935442", "CENTRE DE FORMATION PRO", REFERENCE, "", "BAF"],
    ["PROFESSIONAL TRAINING CENTER BAFOUSSAM", DATE_DEFAUT, "672011135", "CENTRE DE FORMATION PRO", REFERENCE, "", "BAF"],
    ["ISTAMA DOUALA", DATE_DEFAUT, "233421698/ 233427330", "INSTITUT SUPERIEUR", REFERENCE, "", "DLA"],
    ["CFP WESCO CAMEROON", DATE_DEFAUT, "698512623", "CENTRE DE FORMATION PRO", REFERENCE, "", "DLA"],
    ["SEVEN ADVANCED ACADEMY", DATE_DEFAUT, "658066982", "CENTRE DE FORMATION DIGITALE", REFERENCE, "", "DLA"],
    ["LOCALHOST ACADEMY", DATE_DEFAUT, "678279957", "CENTRE DE FORMATION DIGITALE", REFERENCE, "", "DLA"],
    ["CODINGHQ", DATE_DEFAUT, "6819011252", "CENTRE DE FORMATION DIGITALE", REFERENCE, "", "DLA"],
    ["KOMEPE", DATE_DEFAUT, "690131340", "CENTRE DE FORMATION PRO", REFERENCE, "", "YDÉ, DLA"],
    ["CFPY YAMDJEU", DATE_DEFAUT, "699642501", "CENTRE DE FORMATION PRO", REFERENCE, "", "YDÉ"],
    ["CFP MESAKA", DATE_DEFAUT, "691091307", "CENTRE DE FORMATION PRO", REFERENCE, "", "CM"],
    ["INSTITUT VISA", DATE_DEFAUT, "675445324", "CENTRE DE FORMATION PRO", REFERENCE, "", "YDÉ"],
    # --- Beauté / mode / couture ---
    ["ACADEMIE INTERNATIONALE DE BEAUTE (AIB)", DATE_DEFAUT, "233437450/ 675136036", "ECOLE BEAUTE / ESTHETIQUE", REFERENCE, "", "DLA"],
    ["INSTITUT FKENG MODE", DATE_DEFAUT, "699780353/ 693209976/ 679179032", "ECOLE MODE / BEAUTE", REFERENCE, "", "YDÉ"],
    ["CAMAL HAUTE COUTURE", DATE_DEFAUT, "694227924", "ECOLE COUTURE / MODE", REFERENCE, "", "DLA"],
    # --- Auto-écoles / conduite ---
    ["AUTO ECOLE MIRA", DATE_DEFAUT, "699949952/ 695070586", "AUTO ECOLE", REFERENCE, "", "DLA"],
    ["AUTO ECOLE MONTHE / CFP GROUPE MONTHE", DATE_DEFAUT, "675218871/ 691039742", "AUTO ECOLE / FORMATION MONITEURS", REFERENCE, "", "YDÉ, DLA, BAF"],
    # --- Formateurs indépendants / coachs ---
    ["AURELE SIMO (GRIOTYS ACADEMY)", DATE_DEFAUT, "699067626", "FORMATEUR INDEPENDANT", REFERENCE, "", "YDÉ, DLA"],
    ["SIBAFO DAVID (DIGITAL DOCTA)", DATE_DEFAUT, "678532388/ 693077332", "FORMATEUR INDEPENDANT", REFERENCE, "", "CM"],
    # --- Autres structures formation ---
    ["INSTITUT CATHOLIQUE DE BAFOUSSAM (ICABAF)", DATE_DEFAUT, "242000350", "INSTITUT SUPERIEUR", REFERENCE, "", "BAF"],
    ["CENTRE TECHNIQUE DE GAROUA", DATE_DEFAUT, "699500270", "CENTRE DE FORMATION PRO", REFERENCE, "", "GAR"],
    ["CFP PARAMEDICAL AREWA PLUS", DATE_DEFAUT, "674343320", "CENTRE DE FORMATION SANTE", REFERENCE, "", "GAR"],
    # --- Hôtellerie / cuisine ---
    ["SESAME HOTELLERIE", DATE_DEFAUT, "690234576/ 679197725", "ECOLE HOTELLERIE / CUISINE", REFERENCE, "", "YDÉ"],
    # --- GoAfrica / autres ---
    ["CREFPEH OBILI", DATE_DEFAUT, "222318354", "CENTRE DE FORMATION PRO", REFERENCE, "", "YDÉ"],
    ["CFP TRANSFORMING OUR WORLD", DATE_DEFAUT, "693879647", "CENTRE DE FORMATION PRO", REFERENCE, "", "YDÉ"],
    ["NEW-AFRICA METIERS DU BATIMENT", DATE_DEFAUT, "683257782", "CENTRE DE FORMATION PRO", REFERENCE, "", "YDÉ"],
    ["INSTITUT PROFESSIONNEL DE CERTIFICATION", DATE_DEFAUT, "691289564", "CENTRE DE FORMATION PRO", REFERENCE, "", "YDÉ"],
    ["CENTRE CONDUITE ENGINS LOURDS NSAM", DATE_DEFAUT, "675976788", "CENTRE DE FORMATION PRO", REFERENCE, "", "YDÉ"],
]


def normalize_phone(raw: str) -> str:
    if not raw:
        return ""
    parts = re.split(r"[/,;]| et ", raw)
    cleaned = []
    for p in parts:
        digits = re.sub(r"\D", "", p)
        if digits.startswith("237") and len(digits) >= 12:
            digits = digits[3:]
        if len(digits) >= 8:
            cleaned.append(digits)
    return "/ ".join(dict.fromkeys(cleaned))


# Dédupliquer par téléphone principal
seen = set()
final_rows = []
for nom, date, phone, typ, ref, retour, ville in ROWS:
    phone_n = normalize_phone(phone)
    key = re.sub(r"\D", "", phone_n)[:9]
    if key in seen:
        continue
    seen.add(key)
    final_rows.append([nom.upper(), date, phone_n, typ, ref, retour, ville])

wb = Workbook()
ws = wb.active
ws.title = "Prospects formation"

header_fill = PatternFill("solid", fgColor="0F3D2E")
header_font = Font(bold=True, color="FFFFFF")
thin = Border(
    left=Side(style="thin", color="D0D5DD"),
    right=Side(style="thin", color="D0D5DD"),
    top=Side(style="thin", color="D0D5DD"),
    bottom=Side(style="thin", color="D0D5DD"),
)

ws.append(HEADERS)
for c in range(1, 8):
    cell = ws.cell(1, c)
    cell.fill = header_fill
    cell.font = header_font
    cell.border = thin
    cell.alignment = Alignment(vertical="center", wrap_text=True)

for row in final_rows:
    ws.append(row)
    r = ws.max_row
    for c in range(1, 8):
        cell = ws.cell(r, c)
        cell.border = thin
        cell.alignment = Alignment(vertical="center", wrap_text=True)

widths = [42, 12, 34, 32, 12, 22, 18]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

ws.freeze_panes = "A2"
ws.auto_filter.ref = f"A1:G{ws.max_row}"

out = r"C:\Users\zidane\Nexa\academie-langues\docs\prospection\nexa-prospects-formation-elargie.xlsx"
wb.save(out)
print(out)
print(f"{len(final_rows)} lignes")
