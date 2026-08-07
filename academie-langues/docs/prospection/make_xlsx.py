# -*- coding: utf-8 -*-
"""Génère nexa-prospects-b2b.xlsx au format colonnes utilisateur."""
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import re

HEADERS = [
    "NOM ETS",
    "DATE",
    "NO TELEPHONE",
    "TYPE DE STRUCTURE",
    "REFERENCE",
    "RETOUR APRES  RELANCE",
    "VILLE",
]

REFERENCE = "ZIDANE"
DATE_DEFAUT = "03/08/2026"


def normalize_phone(raw: str) -> str:
    if not raw:
        return ""
    # garder chiffres et /
    parts = re.split(r"[/,;]| et ", raw)
    cleaned = []
    for p in parts:
        digits = re.sub(r"\D", "", p)
        if digits.startswith("237") and len(digits) >= 12:
            digits = digits[3:]
        if digits.startswith("225") and len(digits) > 9:
            digits = digits  # garder indicatif international hors CM
        if digits.startswith("221") or digits.startswith("1877") or digits.startswith("1"):
            cleaned.append("+" + digits if not digits.startswith("+") else digits)
            continue
        if "@" in p:
            continue
        if len(digits) >= 8:
            cleaned.append(digits)
    # emails seuls
    if not cleaned and "@" in raw:
        return raw.strip()
    return "/ ".join(dict.fromkeys(cleaned))  # unique order-preserving


def ville_code(ville: str) -> str:
    v = (ville or "").upper().replace("É", "E").replace("É", "E")
    mapping = [
        ("YAOUND", "YDÉ"),
        ("DOUALA", "DLA"),
        ("BAFOUSSAM", "BAF"),
        ("DSCHANG", "DSG"),
        ("ABIDJAN", "ABJ"),
        ("DAKAR", "DKR"),
        ("ABBOTSFORD", "ABB"),
        ("MARYLAND", "MD"),
    ]
    out = []
    for name, code in mapping:
        if name in v and code not in out:
            out.append(code)
    return ", ".join(out)


def type_structure(typ: str) -> str:
    t = (typ or "").lower()
    if any(k in t for k in ("digital", "informatique", "cyber", "web mobile", "metiers du digital")):
        return "CENTRE DE FORMATION DIGITALE"
    if any(k in t for k in ("professionnelle", "cfp", "formation pro")) and "langue" not in t:
        return "CENTRE DE FORMATION PRO"
    if any(k in t for k in ("tcf", "tef", "ielts", "toefl", "langue", "linguist", "allemand", "anglais", "goethe", "chinois", "italien")):
        return "CENTRE LINGUSTIQUE"
    return "CENTRE LINGUSTIQUE"


# Exemples fournis par l'utilisateur (en tête)
ROWS = [
    ["TD BRIDGE INSTITUT", "03/08/2026", "691421692/ 692245973/ 698483258", "CENTRE LINGUSTIQUE", "ZIDANE", "", "YDÉ, DLA, BAF"],
    ["", "03/08/2026", "682594788", "CENTRE LINGUSTIQUE", "ZIDANE", "", "YDÉ"],
    ["INSTITUT N.K", "", "686316976", "CENTRE LINGUSTIQUE", "ZIDANE", "", "YDÉ"],
    ["INSTITUT GHOMS", "", "693071789", "CENTRE LINGUSTIQUE", "ZIDANE", "", "YDÉ"],
]

# Seed collecté (nom, type_formation, contact, pays, ville)
SEED = [
    ("Kennis Language Center", "Ecole de langues", "+237650309085", "Cameroun", "Douala / Yaounde"),
    ("Fishboard", "Preparation IELTS anglais", "+237656525753", "Cameroun", "Yaounde"),
    ("GTIC ecole de langues", "Langues allemand anglais", "+237675723262", "Cameroun", "Douala"),
    ("Seven Advanced Academy", "Formation digitale", "+237658066982", "Cameroun", "Douala"),
    ("PSK Language Center", "Ecole de langues anglais", "+237692351189", "Cameroun", "Douala"),
    ("English For All Cameroun", "Centre linguistique anglais", "+237699794868", "Cameroun", "Yaounde"),
    ("VLC Vista Language Center", "Langues TOEIC TOEFL IELTS", "+237695223236", "Cameroun", "Douala"),
    ("LocalHost Academy", "Metiers du digital", "+237678279957", "Cameroun", "Douala"),
    ("Bamacours Douala", "Langues formation pro", "+237657435416", "Cameroun", "Douala"),
    ("Bamacours Abidjan", "Langues formation pro B2B", "+2250720719706", "Cote d'Ivoire", "Abidjan"),
    ("Bamacours Dakar", "Langues formation pro", "+221785971415", "Senegal", "Dakar"),
    ("Institut Mozart", "Langues IELTS TOEFL TCF Goethe", "+237676365543", "Cameroun", "Yaounde"),
    ("ICLY Yaounde", "Langues DE IT FR EN", "+237699674759", "Cameroun", "Yaounde"),
    ("Erfolg-Zentrum", "Allemand intensif Goethe", "+237699072798", "Cameroun", "Douala"),
    ("Dialog-Haus", "Ecole de langues", "+237674008921", "Cameroun", "Yaounde / Douala"),
    ("Centre Linguistique Chinois", "Chinois mandarin HSK", "+237696295622", "Cameroun", "Douala"),
    ("Solution Centre de Linguistique", "Anglais allemand TOEFL IELTS TCF", "+237621521435", "Cameroun", "Douala"),
    ("Churchill Training Institute CTI", "EN DE IT FR TOEFL IELTS Goethe", "+237673919490", "Cameroun", "Douala"),
    ("Center Multilinguistique Sans Frontiere", "Ecole de langues", "+237696451032", "Cameroun", "Douala"),
    ("Visiocare Academy", "Langues IELTS immigration", "+237691179127", "Cameroun", "Douala"),
    ("Centre Linguistique Visiocare", "Langues internationales", "+237696492222", "Cameroun", "Douala"),
    ("M-Travel cours de langues", "5 langues 4 villes", "+237692904870", "Cameroun", "Yaounde / Douala / Bafoussam"),
    ("Africol", "Anglais IELTS TOEFL DELF TCF", "+237697450252", "Cameroun", "Bafoussam"),
    ("AER Bafoussam", "Allemand Goethe", "+237698224060", "Cameroun", "Bafoussam"),
    ("AER Yaounde", "Allemand Goethe", "+237677451169", "Cameroun", "Yaounde"),
    ("AER Douala", "Allemand Goethe", "+237678196087", "Cameroun", "Douala"),
    ("Prepa TCF Canada C2 Assures", "Preparation TCF Canada", "+237693202340", "Cameroun", "Yaounde / Douala"),
    ("CFP WESCO Cameroon", "Langues + formation pro", "+237698512623", "Cameroun", "Douala"),
    ("American Language Center Douala", "Anglais", "+237675923507", "Cameroun", "Douala"),
    ("British Council Cameroun", "Anglais examens", "+237675291390", "Cameroun", "Yaounde"),
    ("LANGUAGE SERVICE ACADEMY", "Ecole de langues", "+237696206731", "Cameroun", "Douala"),
    ("NEW COUNTRY NEW LANGUAGE", "Ecole de langues", "+237678970572", "Cameroun", "Douala"),
    ("Peerless Language Academy", "Ecole de langues", "+237698296756", "Cameroun", "Douala"),
    ("SLZ-DOUALA", "Allemand Sprachschule", "+237699933738", "Cameroun", "Douala"),
    ("AKADEMIA YAOUNDE", "Ecole de langues", "+237679224360", "Cameroun", "Yaounde"),
    ("Institut Visa", "Langues + visa", "+237675445324", "Cameroun", "Yaounde"),
    ("Clevertrans Dschang", "Ecole de langues", "+237677393931", "Cameroun", "Dschang"),
    ("Africa Services Bafoussam", "Ecole de langues", "+237698625219", "Cameroun", "Bafoussam"),
    ("Bristols House", "Francais anglais mobilite", "+237672304364", "Cameroun", "Douala"),
    ("Cali Pimo Italien", "Italien", "+237657518690", "Cameroun", "Yaounde"),
    ("Carinthon Translation", "Langues traduction", "+237696481325", "Cameroun", "Douala"),
    ("Cenpel Makepe", "Ecole de langues", "+237657095932", "Cameroun", "Douala"),
    ("Der Begleiter Dschang", "Allemand", "+237675427940", "Cameroun", "Dschang"),
    ("Der Begleiter Bafoussam", "Allemand", "+237699161895", "Cameroun", "Bafoussam"),
    ("Center Of Student", "Ecole de langues", "+237698080917", "Cameroun", "Yaounde"),
    ("Cfp Gti Academy", "Ecole de langues", "+237698393939", "Cameroun", "Yaounde"),
    ("Anglais Carrieres Essos", "Anglais bilinguisme pro", "+237671991021", "Cameroun", "Yaounde"),
    ("Das Studentenzentrum", "Ecole de langues", "+237699164974", "Cameroun", "Yaounde"),
    ("Der Erfolg Bepanda", "Allemand anglais", "+237678795275", "Cameroun", "Douala"),
    ("CIS Formation Douala", "Informatique web mobile", "+237698482438", "Cameroun", "Douala"),
    ("CIS Formation Yaounde", "Informatique web mobile", "+237671049520", "Cameroun", "Yaounde"),
    ("KOMEPE", "Formation professionnelle", "+237690131340", "Cameroun", "Douala / Yaounde"),
    ("CFP La Vocation", "Formation professionnelle", "+237657491158", "Cameroun", "Yaounde"),
    ("CFPY Yamdjeu", "Formation professionnelle", "+237699642501", "Cameroun", "Yaounde"),
    ("CFP Mesaka", "Formation professionnelle", "+237691091307", "Cameroun", "Cameroun"),
    ("Crack the TCF Canada", "Preparation TCF TEF", "+18775131563", "Canada", "Abbotsford"),
]

seen_phones = set()
for r in ROWS:
    for d in re.findall(r"\d{8,}", r[2]):
        seen_phones.add(d[-9:] if len(d) >= 9 else d)

for nom, typ, contact, pays, ville in SEED:
    phone = normalize_phone(contact)
    if not phone:
        continue
    key = re.sub(r"\D", "", phone)[:9]
    if key in seen_phones:
        continue
    seen_phones.add(key)
    ROWS.append([
        nom.upper(),
        DATE_DEFAUT,
        phone,
        type_structure(typ),
        REFERENCE,
        "",
        ville_code(ville) or ville.upper(),
    ])

wb = Workbook()
ws = wb.active
ws.title = "Prospects"

header_fill = PatternFill("solid", fgColor="11224E")
header_font = Font(bold=True, color="FFFFFF")
thin = Border(
    left=Side(style="thin", color="D0D5DD"),
    right=Side(style="thin", color="D0D5DD"),
    top=Side(style="thin", color="D0D5DD"),
    bottom=Side(style="thin", color="D0D5DD"),
)

ws.append(HEADERS)
for c in range(1, 8):
    cell = ws.cell(1, c)
    cell.fill = header_fill
    cell.font = header_font
    cell.border = thin
    cell.alignment = Alignment(vertical="center", wrap_text=True)

for row in ROWS:
    ws.append(row)
    r = ws.max_row
    for c in range(1, 8):
        cell = ws.cell(r, c)
        cell.border = thin
        cell.alignment = Alignment(vertical="center", wrap_text=True)

widths = [34, 12, 32, 28, 12, 22, 18]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

ws.freeze_panes = "A2"
ws.auto_filter.ref = f"A1:G{ws.max_row}"

out = r"C:\Users\zidane\Nexa\academie-langues\docs\prospection\nexa-prospects-b2b.xlsx"
wb.save(out)
print(out)
print(f"{len(ROWS)} lignes")
